"""
Parsea Veredas.2 Prueba.xlsx y genera los JSON en /data.
Genera datos sintéticos para abril-mayo 2026.
Run: python3 scripts/seed-from-xlsx.py /path/to/xlsx
"""
import json
import sys
import random
from datetime import datetime, timedelta
from pathlib import Path
import openpyxl

ROOT = Path(__file__).resolve().parents[1]
DATA = ROOT / "data"

CUADRILLAS = ["Adrian", "Mario", "Tyson", "Matias"]

MATERIAL_FIELDS = [
    ("caños", "Caños"),
    ("bolson", "Bolson"),
    ("volquetes", "Volquetes"),
    ("palletBaldosas", "Pallet Baldosas"),
    ("arena", "Arena"),
    ("cal", "Cal"),
    ("cemento", "Cemento"),
    ("mlPlantera", "Ml Plantera"),
    ("protectorPluvial", "Protector Pluvial"),
    ("tapasCamaraPluvial", "Tapas con camara pluvial"),
    ("cordon", "Cordon"),
]

def to_date(v):
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, str) and v:
        try:
            return datetime.fromisoformat(v).strftime("%Y-%m-%d")
        except Exception:
            return None
    return None

def num(v, default=0):
    if v is None or v == "" or v == "#N/A":
        return default
    try:
        return float(v)
    except Exception:
        return default

def parse_rdos(ws):
    headers = [c.value for c in ws[1]]
    rows = []
    for r in range(2, ws.max_row + 1):
        row = {h: ws.cell(row=r, column=i + 1).value for i, h in enumerate(headers) if h}
        fecha = to_date(row.get("Fecha"))
        frente = row.get("Frente")
        if not fecha or not frente:
            continue
        rows.append({
            "id": f"rdo-{r:04d}",
            "fecha": fecha,
            "frente": str(frente).strip(),
            "cuadrilla": str(row.get("Cuadrilla") or "").strip(),
            "m2": num(row.get("M² Ejecutados")),
            "trabajo": "Hormigon" if str(row.get("Trabajo") or "").lower().startswith("hormigon") else "Baldosas",
            "materiales": {
                key: num(row.get(label)) for key, label in MATERIAL_FIELDS
            },
        })
    return rows

def synth_rdos(real_rdos, months_to_add=2):
    """Genera RDOs sintéticos para abril y mayo 2026 manteniendo la distribución."""
    if not real_rdos:
        return []
    random.seed(42)
    frentes = list({r["frente"] for r in real_rdos})
    avg_m2 = sum(r["m2"] for r in real_rdos) / len(real_rdos)
    out = []
    rid = len(real_rdos) + 1
    base_months = [datetime(2026, 4, 1), datetime(2026, 5, 1)]
    nuevas_calles = ["Cabildo", "Olleros", "Echeverria", "Vidal", "Crisologo Larralde", "Juramento", "Manzanares", "Cuba", "Migueletes", "Galvan", "Iberá", "Ciudad de la Paz"]
    for month_start in base_months[:months_to_add]:
        for day in range(15):
            date = month_start + timedelta(days=day * 2)
            cuadrilla = random.choice(CUADRILLAS)
            m2 = round(random.uniform(avg_m2 * 0.5, avg_m2 * 1.5), 1)
            calle = random.choice(nuevas_calles)
            altura = random.randint(2000, 4900)
            frente = f"{calle} {altura}"
            out.append({
                "id": f"rdo-{rid:04d}",
                "fecha": date.strftime("%Y-%m-%d"),
                "frente": frente,
                "cuadrilla": cuadrilla,
                "m2": m2,
                "trabajo": "Baldosas" if random.random() > 0.15 else "Hormigon",
                "materiales": {
                    "caños": 0,
                    "bolson": round(random.uniform(0, 3), 1),
                    "volquetes": max(1, round(m2 / 20)),
                    "palletBaldosas": round(m2 / 25),
                    "arena": round(random.uniform(1, 4)),
                    "cal": round(m2 / 5),
                    "cemento": round(m2 / 5),
                    "mlPlantera": round(random.uniform(0, 12), 1),
                    "protectorPluvial": random.randint(0, 3),
                    "tapasCamaraPluvial": random.randint(0, 2),
                    "cordon": round(random.uniform(0, 8), 1),
                },
            })
            rid += 1
    return out

def parse_precios_cuadrilla(ws):
    out = []
    for r in range(3, 7):
        cuadrilla = ws.cell(row=r, column=2).value
        fecha = to_date(ws.cell(row=r, column=3).value)
        precio = num(ws.cell(row=r, column=4).value, default=1)
        if cuadrilla and fecha:
            out.append({
                "cuadrilla": cuadrilla,
                "vigenciaDesde": fecha,
                "precioM2": precio,
            })
    return out

def parse_precios_material(ws):
    out = []
    unidad_map = {
        "Baldosas": "m²", "Volquetes": "m²", "Caños": "u", "Arena": "u",
        "Cal": "u", "Cemento": "u", "Polietileno": "u", "Seña Pallet": "u", "Bolson": "u",
    }
    for r in range(3, ws.max_row + 1):
        material = ws.cell(row=r, column=2).value
        fecha = to_date(ws.cell(row=r, column=3).value)
        precio = num(ws.cell(row=r, column=4).value, default=1)
        if material and fecha:
            out.append({
                "material": material,
                "vigenciaDesde": fecha,
                "precio": precio,
                "unidad": unidad_map.get(material, "u"),
            })
    return out

MES_MAP = {
    "ENERO": "2026-01", "FEBRERO": "2026-02", "MARZO": "2026-03",
    "ABRIL": "2026-04", "MAYO": "2026-05", "JUNIO": "2026-06",
    "JULIO": "2026-07", "AGOSTO": "2026-08", "SEPTIEMBRE": "2026-09",
    "OCTUBRE": "2026-10", "NOVIEMBRE": "2026-11", "DICIEMBRE": "2026-12",
    "ENERO2": "2027-01", "FEBRERO3": "2027-02",
}

def parse_egresos(ws):
    headers = [ws.cell(row=2, column=c).value for c in range(3, 17)]
    meses_iso = [MES_MAP.get(str(h).strip().upper()) if h else None for h in headers]
    egresos_por_mes = {m: {"impuestos": [], "planesPago": [], "costosFijos": [], "costosVariablesEstimados": 0} for m in meses_iso if m}

    # Impuestos y planes de pago: rows 3-7
    for r in range(3, 8):
        concepto = ws.cell(row=r, column=1).value
        venc = ws.cell(row=r, column=2).value
        if not concepto or "TOTAL" in str(concepto).upper():
            continue
        for ci, mes in enumerate(meses_iso):
            if not mes:
                continue
            v = num(ws.cell(row=r, column=3 + ci).value)
            if v > 0:
                bucket = "planesPago" if str(concepto).startswith("PP.") else "impuestos"
                egresos_por_mes[mes][bucket].append({
                    "concepto": str(concepto).strip(),
                    "vencimiento": str(venc) if venc else None,
                    "monto": v,
                })

    # Costos fijos: rows 12-15 (skip TOTAL)
    for r in range(12, 16):
        concepto = ws.cell(row=r, column=1).value
        venc = ws.cell(row=r, column=2).value
        if not concepto or "TOTAL" in str(concepto).upper():
            continue
        for ci, mes in enumerate(meses_iso):
            if not mes:
                continue
            v = num(ws.cell(row=r, column=3 + ci).value)
            if v > 0:
                egresos_por_mes[mes]["costosFijos"].append({
                    "concepto": str(concepto).strip(),
                    "vencimiento": str(venc) if venc else None,
                    "monto": v,
                })

    # Costos variables: agregar el total mensual
    for ci, mes in enumerate(meses_iso):
        if not mes:
            continue
        # row 26 has TOTAL A PAGAR de variables (sum of rows 21-25)
        total_var = num(ws.cell(row=26, column=3 + ci).value)
        egresos_por_mes[mes]["costosVariablesEstimados"] = total_var

    return [
        {"mes": m, **data}
        for m, data in sorted(egresos_por_mes.items())
        if data["impuestos"] or data["planesPago"] or data["costosFijos"]
    ]

def main():
    if len(sys.argv) < 2:
        print("Usage: seed-from-xlsx.py /path/to/xlsx")
        sys.exit(1)
    src = Path(sys.argv[1]).resolve()
    if not src.exists():
        print(f"Not found: {src}")
        sys.exit(1)

    wb = openpyxl.load_workbook(src, data_only=True)

    rdos_real = parse_rdos(wb["RDO"])
    rdos_synth = synth_rdos(rdos_real)
    rdos = rdos_real + rdos_synth
    print(f"  RDOs real: {len(rdos_real)}, sintéticos: {len(rdos_synth)}, total: {len(rdos)}")

    precios_cuadrilla = parse_precios_cuadrilla(wb["COSTOS"])
    print(f"  Precios cuadrilla: {len(precios_cuadrilla)}")

    precios_material = parse_precios_material(wb["Hoja13"])
    print(f"  Precios material: {len(precios_material)}")

    egresos = parse_egresos(wb["Outs X Mes"])
    print(f"  Egresos: {len(egresos)} meses")

    DATA.mkdir(exist_ok=True)
    (DATA / "rdos.json").write_text(json.dumps(rdos, indent=2, ensure_ascii=False))
    (DATA / "precios-cuadrilla.json").write_text(json.dumps(precios_cuadrilla, indent=2, ensure_ascii=False))
    (DATA / "precios-material.json").write_text(json.dumps(precios_material, indent=2, ensure_ascii=False))
    (DATA / "egresos.json").write_text(json.dumps(egresos, indent=2, ensure_ascii=False))
    print(f"OK JSONs escritos en {DATA}")

if __name__ == "__main__":
    main()
